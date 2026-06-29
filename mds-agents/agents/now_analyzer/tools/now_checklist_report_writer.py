"""Tool for writing NOW application completeness findings to Excel.

Generates an Excel report matching the now_regional_checklist.xlsx format
with checkbox status in column A, plus page numbers and summary notes.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any, Optional

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError as e:
    load_workbook = None
    _IMPORT_ERROR = str(e)

_TEMPLATE_PATH = Path(__file__).with_name("now_regional_checklist.xlsx")
_OUTPUT_DIR = Path.cwd() / "reports"


def write_now_checklist_report(
    findings: list[dict[str, Any]],
    output_filename: str = "now_application_completeness_report.xlsx",
    applicant_name: Optional[str] = None,
    now_number: Optional[str] = None,
) -> str:
    """Write NOW application completeness findings to Excel report.

    Maps findings to the actual now_regional_checklist.xlsx template structure,
    filling in status, evidence, file names, and page numbers for each requirement row.

    Args:
        findings: List of finding records with keys:
            - section: Section name from checklist
            - requirement: Requirement text from checklist
            - status: 'True', 'False', or 'N/A'
            - actual_value: Concrete extracted value (optional, e.g., "Bridget Tetarenko")
            - summary: One-sentence summary (optional, for backward compatibility)
            - evidence_value: Actual evidence found (optional, short text/value)
            - file_name: Name of file containing evidence (optional)
            - page_number: Page number where evidence found (optional)
            - is_instruction: Boolean; True if item is an instruction header
        output_filename: Name of output Excel file.
        applicant_name: Optional applicant/project name.
        now_number: Optional NOW application number.

    Returns:
        Path to generated Excel file as string.

    Raises:
        ImportError: If openpyxl is not available.
        ValueError: If findings data is invalid.
    """
    if load_workbook is None:
        diagnostic = (
            f"openpyxl is required for Excel report generation but not found.\\n"
            f"Python executable: {sys.executable}\\n"
            f"Python version: {sys.version}\\n"
            f"Import error: {_IMPORT_ERROR}\\n\\n"
            f"Fix: Run 'pip install openpyxl' in the environment at {sys.executable}"
        )
        raise ImportError(diagnostic)

    _OUTPUT_DIR.mkdir(exist_ok=True)
    output_path = _OUTPUT_DIR / output_filename

    try:
        wb = load_workbook(str(_TEMPLATE_PATH))
        ws = wb.active
    except Exception as exc:
        raise ValueError(f"Failed to load template workbook: {exc}") from exc

    # Build a map of (section, requirement) -> row_number from template
    # This scans the template to find where each checklist item is located
    template_map: dict[tuple[str, str], int] = {}
    current_section = ""
    
    for row_num in range(1, ws.max_row + 1):
        col_a = ws.cell(row=row_num, column=1).value
        col_c = ws.cell(row=row_num, column=3).value
        
        # Update current section if column A has text
        if col_a and isinstance(col_a, str) and col_a.strip():
            section_text = col_a.strip()
            # Only update section if it looks like a section header (not a checkbox marker)
            if section_text not in ("Required", "Maybe Required", "Missing", "☐"):
                current_section = section_text
        
        # If column C has text, this is a requirement row
        if col_c and isinstance(col_c, str) and col_c.strip() and current_section:
            req_text = col_c.strip()
            template_map[(current_section, req_text)] = row_num

    # Apply applicant info to cells if provided
    if applicant_name:
        try:
            ws["B2"].value = applicant_name
        except:
            pass
    if now_number:
        try:
            ws["B3"].value = now_number
        except:
            pass

    # Define cell styles
    instruction_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    instruction_font = Font(bold=True, color="FFFFFF", size=11)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Process findings and update template rows
    for finding in findings:
        if not isinstance(finding, dict):
            continue

        section = finding.get("section", "").strip()
        requirement = finding.get("requirement", "").strip()
        status = finding.get("status", "").strip()
        actual_value = finding.get("actual_value", "").strip()
        summary = finding.get("summary", "").strip()
        evidence_value = finding.get("evidence_value", "").strip()
        # Column D should contain only the concrete extracted value.
        # Keep Column E for evidence narrative/context.
        if not actual_value:
            actual_value = evidence_value
        evidence_text = summary or evidence_value
        file_name = finding.get("file_name", "").strip()
        page_number = finding.get("page_number", "").strip()
        is_instruction = finding.get("is_instruction", False)

        if not requirement:
            continue

        # Find the row for this requirement in the template
        template_key = (section, requirement)
        if template_key not in template_map:
            continue  # Skip if not found in template

        row_num = template_map[template_key]
        
        # Determine formatting based on status
        if is_instruction:
            checkbox_value = ""
        elif status == "True":
            checkbox_value = "☑"
        elif status in ("False", "N/A"):
            checkbox_value = "☐"
        else:
            checkbox_value = "☐"

        # Update cells in the template row.
        # Column A: Checkbox status (checked/unchecked)
        col_a = ws.cell(row=row_num, column=1)
        col_a.value = checkbox_value
        col_a.alignment = center_alignment
        col_a.border = thin_border

        # Column B: leave blank; status is represented via checkbox in column A.
        col_b = ws.cell(row=row_num, column=2)
        col_b.value = ""
        col_b.fill = instruction_fill if is_instruction else PatternFill(fill_type=None)
        col_b.font = instruction_font if is_instruction else Font(color="000000", size=10)
        col_b.alignment = center_alignment
        col_b.border = thin_border

        # Column C: Requirement (already in template, but ensure formatting)
        col_c = ws.cell(row=row_num, column=3)
        col_c.fill = instruction_fill if is_instruction else PatternFill(fill_type=None)
        col_c.font = instruction_font if is_instruction else Font(color="000000", size=10)
        col_c.alignment = left_alignment
        col_c.border = thin_border

        # Column D: Actual value found
        col_d = ws.cell(row=row_num, column=4)
        col_d.value = actual_value
        col_d.fill = instruction_fill if is_instruction else PatternFill(fill_type=None)
        col_d.font = instruction_font if is_instruction else Font(color="000000", size=10)
        col_d.alignment = left_alignment
        col_d.border = thin_border

        # Column E: Evidence narrative / context
        col_e = ws.cell(row=row_num, column=5)
        col_e.value = evidence_text
        col_e.fill = instruction_fill if is_instruction else PatternFill(fill_type=None)
        col_e.font = instruction_font if is_instruction else Font(color="000000", size=10)
        col_e.alignment = left_alignment
        col_e.border = thin_border

        # Column F: File name
        col_f = ws.cell(row=row_num, column=6)
        col_f.value = file_name
        col_f.fill = instruction_fill if is_instruction else PatternFill(fill_type=None)
        col_f.font = instruction_font if is_instruction else Font(color="000000", size=10)
        col_f.alignment = left_alignment
        col_f.border = thin_border

        # Column G: Page number
        col_g = ws.cell(row=row_num, column=7)
        col_g.value = page_number
        col_g.fill = instruction_fill if is_instruction else PatternFill(fill_type=None)
        col_g.font = instruction_font if is_instruction else Font(color="000000", size=10)
        col_g.alignment = center_alignment
        col_g.border = thin_border

    try:
        wb.save(str(output_path))
    except Exception as exc:
        raise ValueError(f"Failed to write Excel file: {exc}") from exc

    return str(output_path)


def parse_findings_from_json(findings_json: str) -> list[dict[str, Any]]:
    """Parse findings from JSON string.

    Args:
        findings_json: JSON string with finding records.

    Returns:
        List of parsed finding dictionaries.
    """
    try:
        return json.loads(findings_json)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid findings JSON: {exc}") from exc
